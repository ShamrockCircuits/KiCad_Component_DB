#!/usr/bin/env python3
"""
Convert an Excel workbook (.xlsx) into a single SQLite database file.

Rules:
  - Sheets whose name starts with an underscore ("_") are skipped.
  - A sheet literally named "Index" (case-insensitive) is skipped.
  - Every other sheet becomes a table in the SQLite database, using the
    sheet's first row as column names and all subsequent rows as data.

Usage:
    python3 xlsx_to_sqlite.py input.xlsx [output.sqlite]

If output.sqlite is omitted, it defaults to the input filename with a
.sqlite extension.
"""

import sys
import sqlite3
import argparse
from pathlib import Path

import openpyxl


def quote_ident(name: str) -> str:
    """Safely quote an SQL identifier (table or column name)."""
    return '"' + str(name).replace('"', '""') + '"'


def unique_column_names(raw_names):
    """
    Make column names unique and non-empty. Blank headers become
    'column_N'; duplicate headers get a numeric suffix.
    """
    seen = {}
    result = []
    for i, name in enumerate(raw_names):
        name = str(name).strip() if name is not None else ""
        if not name:
            name = f"column_{i + 1}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        result.append(name)
    return result


def convert(xlsx_path: Path, sqlite_path: Path, verbose: bool = False):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    if sqlite_path.exists():
        sqlite_path.unlink()

    conn = sqlite3.connect(sqlite_path)
    cur = conn.cursor()

    skipped = []
    converted = []

    for sheet_name in wb.sheetnames:
        if sheet_name.startswith("_"):
            skipped.append(sheet_name)
            if verbose:
                print(f"[skip]  {sheet_name} (underscore-prefixed)")
            continue
        if sheet_name.strip().lower() == "index":
            skipped.append(sheet_name)
            if verbose:
                print(f"[skip]  {sheet_name} (index sheet)")
            continue

        if verbose:
            print(f"[sheet] {sheet_name}")

        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)

        try:
            header = next(rows_iter)
        except StopIteration:
            # Completely empty sheet -- nothing to create.
            skipped.append(f"{sheet_name} (empty)")
            if verbose:
                print(f"[skip]  {sheet_name} (empty sheet)")
            continue

        columns = unique_column_names(header)
        quoted_cols = [quote_ident(c) for c in columns]
        table = quote_ident(sheet_name)

        cur.execute(
            f"CREATE TABLE {table} ({', '.join(f'{c} TEXT' for c in quoted_cols)})"
        )

        placeholders = ", ".join("?" for _ in columns)
        insert_sql = f"INSERT INTO {table} ({', '.join(quoted_cols)}) VALUES ({placeholders})"

        # Use the first column (typically an ID like "IPN") to identify
        # each row in verbose feedback, falling back to a row number.
        row_count = 0
        for row in rows_iter:
            # Skip fully blank rows
            if row is None or all(v is None for v in row):
                continue
            # Pad/truncate row to match header length
            row = list(row[: len(columns)]) + [None] * (len(columns) - len(row))
            cur.execute(insert_sql, row)
            row_count += 1

            if verbose:
                label = row[0] if row[0] not in (None, "") else f"row {row_count}"
                print(f"  [{row_count}] {sheet_name}: inserted {label}")

        converted.append((sheet_name, row_count))
        if verbose:
            print(f"[done]  {sheet_name}: {row_count} row(s) inserted\n")

    conn.commit()
    conn.close()

    print(f"Wrote {sqlite_path}")
    print("\nTables created:")
    for name, count in converted:
        print(f"  {name}: {count} row(s)")
    if skipped:
        print("\nSheets skipped:")
        for name in skipped:
            print(f"  {name}")


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("xlsx_path", type=Path, help="Path to the input .xlsx file")
    parser.add_argument(
        "sqlite_path",
        type=Path,
        nargs="?",
        default=None,
        help="Path to the output .sqlite file (default: same name, .sqlite extension)",
    )
    parser.add_argument(
        "-v", "--verbose",
        action="store_true",
        help="Print feedback for each sheet and each row as it is inserted",
    )
    args = parser.parse_args()

    xlsx_path = args.xlsx_path
    sqlite_path = args.sqlite_path or xlsx_path.with_suffix(".sqlite")

    if not xlsx_path.exists():
        sys.exit(f"Error: {xlsx_path} does not exist")

    convert(xlsx_path, sqlite_path, verbose=args.verbose)


if __name__ == "__main__":
    main()
